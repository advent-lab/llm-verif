#!/usr/bin/env python3
"""Parse line/statement coverage reports from UVM_COVERAGE_MODE=line runs and generate an Excel summary.

Sheets:
  1. Line Coverage       — per-iteration RTL statement % + final merged per run
  2. Cumulative Coverage — cumulative (monotonically increasing) statement % per iteration from run.log
  3. RTL Breakdown       — per-RTL-module statement % per iteration + final
  4. Run Info            — API calls, iterations, tokens, context %, termination, driver pipeline
  5. Per-design sheets   — detailed per-DU breakdown per run
"""

import glob
import os
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORK_DIR = os.path.join(os.path.dirname(__file__), "work")

# Maps display label → directory prefix for code coverage runs
CODE_DESIGNS = {
    "alu":                 "cvdp_agentic_alu_uvm_code",
    "memory_scheduler":    "cvdp_agentic_memory_scheduler_uvm_code",
    "rgb_color_space":     "cvdp_agentic_rgb_color_space_conversion_uvm_code",
    "sha1":                "sha1_top_uvm_code",
    "trng":                "trng_top_uvm_code",
}


# ---------------------------------------------------------------------------
# XML coverage parsing
# ---------------------------------------------------------------------------

def parse_xml_coverage_detailed(filepath):
    """Parse an XML coverage report produced by QuestaSim.

    Returns dict:
        total_pct   — weighted statement % across ALL design units
        rtl_pct     — weighted statement % for RTL-only design units
        per_du      — { du_name: { pct, active, hits, is_rtl, rtl_files } }
    """
    result = {
        "total_pct": None,
        "rtl_pct": None,
        "per_du": {},
    }
    if not os.path.exists(filepath):
        return result

    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
    except Exception:
        return result

    total_active = 0
    total_hits = 0
    rtl_active = 0
    rtl_hits = 0

    for du in root.iter("DuData"):
        du_name = du.get("du", "unknown")
        stmts = du.find("statements")
        if stmts is None:
            continue
        active = int(stmts.get("active", 0))
        hits = int(stmts.get("hits", 0))
        pct = float(stmts.get("percent", 0.0))

        # Identify RTL files (source path contains /rtl/)
        src = du.find("sourceTable")
        rtl_files = []
        if src is not None:
            for fm in src.findall("fileMap"):
                path = fm.get("path", "")
                if "/rtl/" in path:
                    rtl_files.append(os.path.basename(path))
        is_rtl = len(rtl_files) > 0

        result["per_du"][du_name] = {
            "pct": pct,
            "active": active,
            "hits": hits,
            "is_rtl": is_rtl,
            "rtl_files": rtl_files,
        }

        total_active += active
        total_hits += hits
        if is_rtl:
            rtl_active += active
            rtl_hits += hits

    if total_active > 0:
        result["total_pct"] = total_hits / total_active * 100
    if rtl_active > 0:
        result["rtl_pct"] = rtl_hits / rtl_active * 100

    return result


# ---------------------------------------------------------------------------
# Run log parsing
# ---------------------------------------------------------------------------

def parse_run_log_code(log_path):
    """Parse run.log from a code-coverage run.

    Returns dict with:
        total_api_calls, total_iterations, final_tokens, final_context_pct,
        context_window, termination_type, termination_reason,
        driver_pipeline_enabled, driver_file,
        run_start, run_end, run_duration_min,
        max_no_progress, max_failures,
        infra_modification_iter, infra_modification_api_call,
        cumulative_per_iter  — { iter_num: cumulative_pct } (from log)
        iter_pct_per_iter    — { iter_num: iteration_pct } (from log)
    """
    info = {
        "total_api_calls": 0,
        "total_iterations": 0,
        "final_tokens": None,
        "final_context_pct": None,
        "context_window": None,
        "termination_type": None,
        "termination_reason": None,
        "driver_pipeline_enabled": False,
        "driver_file": None,
        "run_start": None,
        "run_end": None,
        "run_duration_min": None,
        "max_no_progress": 0,
        "max_failures": 0,
        "infra_modification_iter": None,
        "infra_modification_api_call": None,
        "cumulative_per_iter": {},   # iter_num → cumulative %
        "iter_pct_per_iter": {},     # iter_num → iteration %
    }

    if not os.path.exists(log_path):
        return info

    current_api_call = 0
    current_iter = 1
    # Track pending iter coverage so we can associate it with the right iter
    pending_iter_pct = None

    with open(log_path, "r") as f:
        for line in f:
            clean = re.sub(r"\x1b\[[0-9;]*m", "", line)

            # Timestamps
            ts_match = re.match(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})", clean)
            if ts_match:
                ts = ts_match.group(1)
                if info["run_start"] is None:
                    info["run_start"] = ts
                info["run_end"] = ts

            # API REQUEST line
            m = re.search(
                r"API REQUEST \[API Call #(\d+) \| Iter (\d+) \| "
                r"Cumulative: [\d.]+% \| Last: [\d.]+% \| "
                r"Failures: (\d+) \| No Progress: (\d+)"
                r"(?: \| NoTool: \d+)?"
                r" \| Tokens: ([\d,]+) \(([\d.]+)%\)",
                clean,
            )
            if m:
                api_num = int(m.group(1))
                iter_num = int(m.group(2))
                failures = int(m.group(3))
                no_progress = int(m.group(4))
                tokens = int(m.group(5).replace(",", ""))
                ctx_pct = float(m.group(6))

                current_api_call = api_num
                current_iter = iter_num

                info["total_api_calls"] = max(info["total_api_calls"], api_num + 1)
                info["total_iterations"] = max(info["total_iterations"], iter_num)
                info["final_tokens"] = tokens
                info["final_context_pct"] = ctx_pct
                info["max_no_progress"] = max(info["max_no_progress"], no_progress)
                info["max_failures"] = max(info["max_failures"], failures)

                if info["context_window"] is None and ctx_pct > 0:
                    info["context_window"] = round(tokens / (ctx_pct / 100))

            # Per-iteration and cumulative coverage lines in run.log
            m_iter = re.search(r"Iteration coverage:\s+([\d.]+)%", clean)
            if m_iter:
                pending_iter_pct = float(m_iter.group(1))
                # Associate with current_iter (the iter that just simulated)
                info["iter_pct_per_iter"][current_iter] = pending_iter_pct

            m_cum = re.search(r"Cumulative coverage:\s+([\d.]+)%", clean)
            if m_cum:
                cum_val = float(m_cum.group(1))
                info["cumulative_per_iter"][current_iter] = cum_val

            # Termination: accepted signal_done
            if "Accepting signal_done" in clean:
                m2 = re.search(r"Accepting signal_done:\s*(.+)", clean)
                if m2:
                    reason = m2.group(1).strip()
                    info["termination_reason"] = reason
                    if "100%" in reason or "coverage" in reason.lower():
                        info["termination_type"] = "coverage_met"
                    elif "no_progress" in reason.lower():
                        info["termination_type"] = "no_progress"
                    else:
                        info["termination_type"] = "graceful"

            # UVM driver detection
            if "UVM driver auto-detected" in clean:
                info["driver_pipeline_enabled"] = True
                m2 = re.search(r"UVM driver auto-detected:\s*(\S+)", clean)
                if m2:
                    info["driver_file"] = m2.group(1)

            # Infrastructure modification
            if ("Infrastructure modification ENABLED" in clean or
                    "infra_modification_granted" in clean or
                    "Infrastructure modification GRANTED" in clean):
                if info["infra_modification_iter"] is None:
                    info["infra_modification_iter"] = current_iter
                    info["infra_modification_api_call"] = current_api_call

            # Summary lines
            m = re.search(r"Iterations:\s+(\d+)", clean)
            if m:
                info["total_iterations"] = max(info["total_iterations"], int(m.group(1)))
            m = re.search(r"API calls:\s+(\d+)", clean)
            if m:
                info["total_api_calls"] = max(info["total_api_calls"], int(m.group(1)))

    # Compute run duration
    if info["run_start"] and info["run_end"]:
        try:
            start = datetime.strptime(info["run_start"], "%Y-%m-%d %H:%M:%S")
            end = datetime.strptime(info["run_end"], "%Y-%m-%d %H:%M:%S")
            info["run_duration_min"] = round((end - start).total_seconds() / 60, 1)
        except ValueError:
            pass

    # Infer termination type if not explicitly found
    if info["termination_type"] is None:
        if info["final_context_pct"] and info["final_context_pct"] >= 90:
            info["termination_type"] = "context_window"
            info["termination_reason"] = "context window exhausted"
        elif info["max_failures"] >= 3:
            info["termination_type"] = "max_retries"
            info["termination_reason"] = "consecutive failures"
        elif info["max_no_progress"] >= 5:
            info["termination_type"] = "max_no_progress"
            info["termination_reason"] = "no progress limit"
        else:
            info["termination_type"] = "max_iterations"
            info["termination_reason"] = "iteration/API call limit"

    return info


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------

def collect_data():
    """Collect line coverage data from all _code_run_N directories."""
    data = {}
    for design_label, design_prefix in CODE_DESIGNS.items():
        data[design_label] = {}
        run_dirs = sorted(glob.glob(os.path.join(WORK_DIR, f"{design_prefix}_run_*")))
        run_nums = []
        for rd in run_dirs:
            m = re.search(r"_run_(\d+)$", rd)
            if m and os.path.isdir(rd):
                run_nums.append(int(m.group(1)))
        run_nums.sort()

        for run_num in run_nums:
            run_dir = os.path.join(WORK_DIR, f"{design_prefix}_run_{run_num}")
            cov_dir = os.path.join(run_dir, "coverage")

            # Discover iteration XML reports
            iter_xml_map = {}
            for xp in glob.glob(os.path.join(cov_dir, "iter_*_report.xml")):
                m = re.search(r"iter_(\d+)_report\.xml", xp)
                if m:
                    iter_xml_map[int(m.group(1))] = xp
            iter_nums = sorted(iter_xml_map.keys())

            # Parse each iteration's XML
            iterations = {}
            for iter_num in iter_nums:
                cov = parse_xml_coverage_detailed(iter_xml_map[iter_num])
                iterations[iter_num] = cov

            # Parse cumulative XML
            cum_xml = os.path.join(cov_dir, "cumulative_report.xml")
            cum_cov = parse_xml_coverage_detailed(cum_xml)

            # Parse run log (includes per-iter cumulative coverage from log)
            log_path = os.path.join(run_dir, "run.log")
            run_info = parse_run_log_code(log_path)

            data[design_label][run_num] = {
                "iterations": iterations,
                "iter_nums": iter_nums,
                "cumulative": cum_cov,
                "run_info": run_info,
            }

    return data


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def create_excel(data, output_path):
    wb = Workbook()

    # ---- Styles ----
    header_font     = Font(bold=True, size=11, color="FFFFFF")
    header_fill     = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    cum_header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    rtl_header_fill = PatternFill(start_color="833C00", end_color="833C00", fill_type="solid")
    info_header_fill= PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    subheader_fill  = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    subheader_font  = Font(bold=True, size=10)
    cum_fill        = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    rtl_fill        = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    pct_format  = "0.00%"
    center      = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_header(ws, row, c1, c2, fill=None):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            cell.font  = header_font
            cell.fill  = fill or header_fill
            cell.alignment = wrap_center
            cell.border = thin_border

    def style_subheader(ws, row, c1, c2):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            cell.font  = subheader_font
            cell.fill  = subheader_fill
            cell.alignment = center
            cell.border = thin_border

    def write_pct(ws, row, col, val, bold=False, fill=None):
        cell = ws.cell(row=row, column=col)
        if val is not None:
            cell.value = val / 100.0
            cell.number_format = pct_format
        else:
            cell.value = "N/A"
        cell.alignment = center
        cell.border = thin_border
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill

    def write_text(ws, row, col, val, bold=False, fill=None, left=False):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.alignment = Alignment(horizontal="left" if left else "center",
                                   vertical="center")
        cell.border = thin_border
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill

    def write_num(ws, row, col, val, fmt=None, bold=False, fill=None):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.alignment = center
        cell.border = thin_border
        if fmt:
            cell.number_format = fmt
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill

    # ------------------------------------------------------------------
    # Compute global max iterations
    # ------------------------------------------------------------------
    max_iters = 0
    for dl in data:
        for rn in data[dl]:
            max_iters = max(max_iters, len(data[dl][rn]["iter_nums"]))

    # ==================================================================
    # Sheet 1: Line Coverage (per-iteration RTL statement %)
    # ==================================================================
    ws = wb.active
    ws.title = "Line Coverage"

    headers = ["Design", "Run"]
    for i in range(1, max_iters + 1):
        headers.append(f"Iter {i}\n(RTL stmt %)")
    headers.append("Final\n(merged RTL %)")
    headers.append("Final\n(merged total %)")
    max_col = len(headers)

    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, 1, max_col, fill=header_fill)
    ws.row_dimensions[1].height = 35

    row = 2
    for design_label in CODE_DESIGNS:
        if design_label not in data:
            continue
        for run_num in sorted(data[design_label]):
            rd      = data[design_label][run_num]
            iters   = rd["iterations"]
            iter_nums = rd["iter_nums"]
            cum     = rd["cumulative"]
            write_text(ws, row, 1, design_label)
            write_text(ws, row, 2, f"run_{run_num}")
            for idx, iter_num in enumerate(iter_nums):
                write_pct(ws, row, 3 + idx, iters[iter_num]["rtl_pct"])
            for i in range(len(iter_nums), max_iters):
                write_text(ws, row, 3 + i, "-")
            write_pct(ws, row, 3 + max_iters,     cum["rtl_pct"],   bold=True, fill=cum_fill)
            write_pct(ws, row, 3 + max_iters + 1, cum["total_pct"], bold=True, fill=cum_fill)
            row += 1

    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 10
    for c in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ==================================================================
    # Sheet 2: Cumulative Coverage (monotonically increasing from log)
    # ==================================================================
    ws = wb.create_sheet(title="Cumulative Coverage")

    headers2 = ["Design", "Run"]
    for i in range(1, max_iters + 1):
        headers2.append(f"After\nIter {i}")
    headers2.append("Final\n(XML merged)")
    max_col2 = len(headers2)

    for ci, h in enumerate(headers2, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, 1, max_col2, fill=cum_header_fill)
    ws.row_dimensions[1].height = 35

    row = 2
    for design_label in CODE_DESIGNS:
        if design_label not in data:
            continue
        for run_num in sorted(data[design_label]):
            rd      = data[design_label][run_num]
            iter_nums = rd["iter_nums"]
            cum     = rd["cumulative"]
            ri      = rd["run_info"]
            cum_per_iter = ri["cumulative_per_iter"]  # iter_num → pct from log

            write_text(ws, row, 1, design_label)
            write_text(ws, row, 2, f"run_{run_num}")
            for idx, iter_num in enumerate(iter_nums):
                val = cum_per_iter.get(iter_num)
                write_pct(ws, row, 3 + idx, val)
            for i in range(len(iter_nums), max_iters):
                write_text(ws, row, 3 + i, "-")
            write_pct(ws, row, 3 + max_iters, cum["rtl_pct"], bold=True, fill=cum_fill)
            row += 1

    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 10
    for c in range(3, max_col2 + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ==================================================================
    # Sheet 3: RTL Module Breakdown (per-run, final merged only)
    # ==================================================================
    ws = wb.create_sheet(title="RTL Breakdown")

    # Gather all RTL DU names across all designs/runs
    all_rtl_dus_by_design = {}
    for dl in CODE_DESIGNS:
        if dl not in data:
            continue
        dus = set()
        for rn in data[dl]:
            for du_name, du_info in data[dl][rn]["cumulative"]["per_du"].items():
                if du_info["is_rtl"]:
                    dus.add(du_name)
        all_rtl_dus_by_design[dl] = sorted(dus)

    # Headers: Design | Run | Final RTL % | <du_1> | <du_2> | ...
    # We'll do it per-design block since DU sets differ
    row = 1
    for design_label in CODE_DESIGNS:
        if design_label not in data:
            continue
        rtl_dus = all_rtl_dus_by_design.get(design_label, [])
        if not rtl_dus:
            continue
        n_cols = 3 + len(rtl_dus)  # Design, Run, Final RTL %, per-DU cols

        # Design header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        ws.cell(row=row, column=1, value=f"{design_label} — RTL Module Coverage")
        style_header(ws, row, 1, n_cols, fill=rtl_header_fill)
        row += 1

        # Column headers
        col_hdrs = ["Design", "Run", "Final\nRTL %"] + [f"{du}\n(stmt %)" for du in rtl_dus]
        for ci, h in enumerate(col_hdrs, 1):
            ws.cell(row=row, column=ci, value=h)
        style_subheader(ws, row, 1, n_cols)
        row += 1

        for run_num in sorted(data[design_label]):
            rd  = data[design_label][run_num]
            cum = rd["cumulative"]
            write_text(ws, row, 1, design_label)
            write_text(ws, row, 2, f"run_{run_num}")
            write_pct(ws, row, 3, cum["rtl_pct"], bold=True, fill=rtl_fill)
            for ci, du_name in enumerate(rtl_dus, 4):
                du_info = cum["per_du"].get(du_name, {})
                write_pct(ws, row, ci, du_info.get("pct"))
            row += 1
        row += 1  # blank row between designs

    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 14
    for c in range(4, 20):
        ws.column_dimensions[get_column_letter(c)].width = 20

    # ==================================================================
    # Sheet 4: Run Info
    # ==================================================================
    ws = wb.create_sheet(title="Run Info")

    info_headers = [
        # Identity
        "Design", "Run",
        # Coverage result
        "Final\nRTL Stmt %", "Final\nTotal Stmt %",
        # Runtime
        "API Calls", "Iterations", "Duration\n(min)",
        # Token usage
        "Final\nTokens", "Context\nUsed %", "Context\nWindow",
        # Termination
        "Termination\nType", "Termination\nReason",
        # Features
        "Driver\nPipeline", "Driver\nFile",
        "Infra Mod\nTriggered", "Infra Mod\n@ Iter", "Infra Mod\n@ API Call",
        # Observed limits
        "Max\nFailures", "Max\nNo-Progress",
        # Timestamps
        "Run Start", "Run End",
    ]
    max_col4 = len(info_headers)
    for ci, h in enumerate(info_headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, 1, max_col4, fill=info_header_fill)
    ws.row_dimensions[1].height = 35

    row = 2
    for design_label in CODE_DESIGNS:
        if design_label not in data:
            continue
        for run_num in sorted(data[design_label]):
            rd  = data[design_label][run_num]
            cum = rd["cumulative"]
            ri  = rd["run_info"]
            c = 1

            write_text(ws, row, c, design_label);                                    c += 1
            write_text(ws, row, c, f"run_{run_num}");                                c += 1
            write_pct(ws,  row, c, cum.get("rtl_pct"));                              c += 1
            write_pct(ws,  row, c, cum.get("total_pct"));                            c += 1
            write_num(ws,  row, c, ri["total_api_calls"]);                           c += 1
            write_num(ws,  row, c, ri["total_iterations"]);                          c += 1
            write_num(ws,  row, c, ri["run_duration_min"], fmt="0.0");               c += 1
            write_num(ws,  row, c, ri["final_tokens"], fmt="#,##0");                 c += 1
            if ri["final_context_pct"] is not None:
                write_pct(ws, row, c, ri["final_context_pct"])
            else:
                write_text(ws, row, c, "N/A")
            c += 1
            write_num(ws,  row, c, ri["context_window"], fmt="#,##0");               c += 1
            write_text(ws, row, c, ri["termination_type"]);                          c += 1
            write_text(ws, row, c, ri["termination_reason"] or "-");                 c += 1
            write_text(ws, row, c, "Yes" if ri["driver_pipeline_enabled"] else "No"); c += 1
            write_text(ws, row, c, ri["driver_file"] or "-");                        c += 1
            write_text(ws, row, c, "Yes" if ri["infra_modification_iter"] else "No"); c += 1
            write_text(ws, row, c, ri["infra_modification_iter"] or "-");            c += 1
            write_text(ws, row, c, ri["infra_modification_api_call"] or "-");        c += 1
            write_num(ws,  row, c, ri["max_failures"]);                              c += 1
            write_num(ws,  row, c, ri["max_no_progress"]);                           c += 1
            write_text(ws, row, c, ri["run_start"] or "-");                          c += 1
            write_text(ws, row, c, ri["run_end"] or "-");                            c += 1

            row += 1

    widths4 = {
        1: 18,  2: 10,  3: 14,  4: 14,
        5: 10,  6: 10,  7: 11,
        8: 12,  9: 11,  10: 12,
        11: 16, 12: 28,
        13: 10, 14: 22, 15: 10, 16: 10, 17: 12,
        18: 10, 19: 12,
        20: 20, 21: 20,
    }
    for col_num, w in widths4.items():
        ws.column_dimensions[get_column_letter(col_num)].width = w

    # ==================================================================
    # Sheet 5+: Per-design detail (all DUs, per-iteration + final)
    # ==================================================================
    for design_label in CODE_DESIGNS:
        if design_label not in data:
            continue
        ws = wb.create_sheet(title=design_label[:31])
        row = 1

        for run_num in sorted(data[design_label]):
            rd        = data[design_label][run_num]
            iters     = rd["iterations"]
            iter_nums = rd["iter_nums"]
            cum       = rd["cumulative"]
            ri        = rd["run_info"]
            n_iters   = len(iter_nums)
            num_cols  = 2 + n_iters + 1  # Metric + iter cols + Final

            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=num_cols)
            ws.cell(row=row, column=1,
                    value=f"Run {run_num}  |  "
                          f"API={ri['total_api_calls']} Iters={ri['total_iterations']} "
                          f"End={ri['termination_type']}")
            style_header(ws, row, 1, num_cols)
            row += 1

            # Sub-header
            write_text(ws, row, 1, "Metric", bold=True)
            for idx, iter_num in enumerate(iter_nums):
                write_text(ws, row, 2 + idx, f"Iter {iter_num}")
            write_text(ws, row, 2 + n_iters, "Final (merged)")
            style_subheader(ws, row, 1, num_cols)
            row += 1

            # RTL-only total
            write_text(ws, row, 1, "RTL Stmt %", bold=True)
            for idx, iter_num in enumerate(iter_nums):
                write_pct(ws, row, 2 + idx, iters[iter_num]["rtl_pct"], fill=rtl_fill)
            write_pct(ws, row, 2 + n_iters, cum["rtl_pct"], bold=True, fill=cum_fill)
            row += 1

            # Total (all DUs)
            write_text(ws, row, 1, "Total Stmt %", bold=True)
            for idx, iter_num in enumerate(iter_nums):
                write_pct(ws, row, 2 + idx, iters[iter_num]["total_pct"])
            write_pct(ws, row, 2 + n_iters, cum["total_pct"], bold=True, fill=cum_fill)
            row += 1

            # Cumulative from log
            cum_per_iter = ri["cumulative_per_iter"]
            write_text(ws, row, 1, "Cumulative % (log)", bold=True)
            for idx, iter_num in enumerate(iter_nums):
                write_pct(ws, row, 2 + idx, cum_per_iter.get(iter_num))
            write_pct(ws, row, 2 + n_iters, cum["rtl_pct"], bold=True, fill=cum_fill)
            row += 1

            # Blank separator
            row += 1

            # Per-DU rows (RTL first, then non-RTL)
            all_du_names = sorted(cum["per_du"].keys(),
                                  key=lambda d: (0 if cum["per_du"][d]["is_rtl"] else 1, d))
            for du_name in all_du_names:
                du_info = cum["per_du"][du_name]
                label = f"  [RTL] {du_name}" if du_info["is_rtl"] else f"  {du_name}"
                fill  = rtl_fill if du_info["is_rtl"] else None
                write_text(ws, row, 1, label, fill=fill)
                for idx, iter_num in enumerate(iter_nums):
                    iter_du = iters[iter_num]["per_du"].get(du_name, {})
                    write_pct(ws, row, 2 + idx, iter_du.get("pct"), fill=fill)
                write_pct(ws, row, 2 + n_iters, du_info.get("pct"),
                          bold=True, fill=cum_fill)
                row += 1

            row += 2  # blank rows between runs

        ws.column_dimensions[get_column_letter(1)].width = 30
        for col in range(2, 20):
            ws.column_dimensions[get_column_letter(col)].width = 16

    output_path = os.path.abspath(output_path)
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    data = collect_data()

    for design in CODE_DESIGNS:
        if design not in data or not data[design]:
            print(f"{design}: no code coverage runs found")
            continue
        for run in sorted(data[design]):
            rd   = data[design][run]
            iters = rd["iterations"]
            iter_nums = rd["iter_nums"]
            cum  = rd["cumulative"]
            ri   = rd["run_info"]

            rtl_parts = []
            for n in iter_nums:
                v = iters[n]["rtl_pct"]
                rtl_parts.append(f"{v:.1f}%" if v is not None else "N/A")
            progression = " → ".join(rtl_parts)
            final_rtl   = cum["rtl_pct"]
            final_total = cum["total_pct"]
            print(
                f"{design} run_{run}: RTL [{progression}] → "
                f"Final RTL={final_rtl:.2f}% Total={final_total:.2f}% | "
                f"API={ri['total_api_calls']} Iters={ri['total_iterations']} "
                f"Tokens={ri['final_tokens']} ({ri['final_context_pct']}%) "
                f"End={ri['termination_type']} "
                f"Driver={'Y' if ri['driver_pipeline_enabled'] else 'N'}"
            )

    create_excel(data, "line_coverage_report.xlsx")
