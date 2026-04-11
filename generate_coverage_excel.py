#!/usr/bin/env python3
"""Parse coverage reports and run logs from work/ directory and generate an Excel summary.

Sheets:
  1. Functional Coverage — per-iteration covergroup % + final merged
  2. Code Coverage — per-iteration statement % + final merged
  3. Run Info — config, token usage, termination, driver pipeline per run
  4. Per-design sheets — detailed covergroup breakdown
"""

import glob
import os
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORK_DIR = os.path.join(os.path.dirname(__file__), "work")

DESIGNS = {
    "alu": "cvdp+agentic_alu_uvm",
    "memory_scheduler": "cvdp+agentic_memory_scheduler_uvm",
    "rgb_color_space": "cvdp+agentic_rgb_color_space_conversion_uvm",
    "sha1": "sha1_top_uvm",
    "trng": "trng_top_uvm",
}


# ---------------------------------------------------------------------------
# Coverage parsing helpers
# ---------------------------------------------------------------------------

def parse_text_summary(filepath):
    """Extract TOTAL COVERGROUP COVERAGE from a text coverage report."""
    covergroup_cov = None
    with open(filepath, "r") as f:
        for line in f:
            m = re.search(r"TOTAL COVERGROUP COVERAGE:\s+([\d.]+)%", line)
            if m:
                covergroup_cov = float(m.group(1))
    return covergroup_cov


def parse_text_covergroup_types_multiline(filepath):
    """Handle TYPE lines where the percentage may be on the next line."""
    cg_pcts = {}
    with open(filepath, "r") as f:
        lines = f.readlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        m = re.match(r"^\s*TYPE\s+\S+/(\w+)\s+([\d.]+)%", line)
        if m:
            name = m.group(1)
            if name not in cg_pcts:
                cg_pcts[name] = float(m.group(2))
            i += 1
            continue
        m = re.match(r"^\s*TYPE\s+\S+/(\w+)\s*$", line)
        if m and i + 1 < len(lines):
            name = m.group(1)
            m2 = re.match(r"^\s+([\d.]+)%", lines[i + 1])
            if m2 and name not in cg_pcts:
                cg_pcts[name] = float(m2.group(1))
            i += 2
            continue
        i += 1
    return cg_pcts


def parse_xml_stmt_coverage(filepath):
    """Extract overall statement coverage % from XML report."""
    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
        total_active = 0
        total_hits = 0
        for du in root.iter("DuData"):
            stmts = du.find("statements")
            if stmts is not None:
                total_active += int(stmts.get("active", 0))
                total_hits += int(stmts.get("hits", 0))
        if total_active > 0:
            return total_hits / total_active * 100
        return None
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Run log parsing
# ---------------------------------------------------------------------------

def parse_run_log(log_path):
    """Parse run.log to extract config, token usage, termination, and events.

    Returns dict with:
        total_api_calls, total_iterations, final_tokens, final_context_pct,
        context_window, termination_reason, termination_type,
        driver_pipeline_enabled, driver_file, run_start, run_end,
        run_duration_min, max_no_progress, max_failures,
        infra_modification_iter, infra_modification_api_call
    """
    info = {
        "total_api_calls": 0,
        "total_iterations": 0,
        "final_tokens": None,
        "final_context_pct": None,
        "context_window": None,
        "termination_reason": None,
        "termination_type": None,  # "graceful" or "limit_reached" or "unknown"
        "driver_pipeline_enabled": False,
        "driver_file": None,
        "run_start": None,
        "run_end": None,
        "run_duration_min": None,
        "max_no_progress": 0,
        "max_failures": 0,
        "infra_modification_iter": None,
        "infra_modification_api_call": None,
    }

    if not os.path.exists(log_path):
        return info

    # Track the current API call/iter context for infra modification detection
    current_api_call = 0
    current_iter = 1

    with open(log_path, "r") as f:
        for line in f:
            # Strip ANSI codes for easier parsing
            clean = re.sub(r"\x1b\[[0-9;]*m", "", line)

            # Timestamps for run duration
            ts_match = re.match(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})", clean)
            if ts_match:
                ts = ts_match.group(1)
                if info["run_start"] is None:
                    info["run_start"] = ts
                info["run_end"] = ts

            # API REQUEST line
            m = re.search(
                r"API REQUEST \[API Call #(\d+) \| Iter (\d+) \| "
                r"Cumulative: [\d.]+% \| Last: [\d.]+% \| "
                r"Failures: (\d+) \| No Progress: (\d+)"
                r"(?: \| NoTool: \d+)?"
                r" \| Tokens: ([\d,]+) \(([\d.]+)%\)",
                clean,
            )
            if m:
                api_num = int(m.group(1))
                iter_num = int(m.group(2))
                failures = int(m.group(3))
                no_progress = int(m.group(4))
                tokens = int(m.group(5).replace(",", ""))
                ctx_pct = float(m.group(6))

                current_api_call = api_num
                current_iter = iter_num

                info["total_api_calls"] = max(info["total_api_calls"], api_num + 1)
                info["total_iterations"] = max(info["total_iterations"], iter_num)
                info["final_tokens"] = tokens
                info["final_context_pct"] = ctx_pct
                info["max_no_progress"] = max(info["max_no_progress"], no_progress)
                info["max_failures"] = max(info["max_failures"], failures)

                # Infer context window from first API call with decent %
                if info["context_window"] is None and ctx_pct > 0:
                    info["context_window"] = round(tokens / (ctx_pct / 100))

            # Termination: accepted signal_done
            if "Accepting signal_done" in clean:
                m2 = re.search(r"Accepting signal_done:\s*(.+)", clean)
                if m2:
                    reason = m2.group(1).strip()
                    info["termination_reason"] = reason
                    if "100%" in reason or "coverage" in reason.lower():
                        info["termination_type"] = "coverage_met"
                    elif "no_progress" in reason:
                        info["termination_type"] = "no_progress"
                    else:
                        info["termination_type"] = "graceful"

            # UVM driver detection
            if "UVM driver auto-detected" in clean:
                info["driver_pipeline_enabled"] = True
                m2 = re.search(r"UVM driver auto-detected:\s*(\S+)", clean)
                if m2:
                    info["driver_file"] = m2.group(1)

            # Infrastructure modification
            if "Infrastructure modification ENABLED" in clean or \
               "infra_modification_granted" in clean or \
               "Infrastructure modification GRANTED" in clean:
                if info["infra_modification_iter"] is None:
                    info["infra_modification_iter"] = current_iter
                    info["infra_modification_api_call"] = current_api_call

            # Detect summary lines at end (from runs that print summary)
            m = re.search(r"Iterations:\s+(\d+)", clean)
            if m:
                info["total_iterations"] = max(info["total_iterations"], int(m.group(1)))
            m = re.search(r"API calls:\s+(\d+)", clean)
            if m:
                info["total_api_calls"] = max(info["total_api_calls"], int(m.group(1)))

    # Compute run duration
    if info["run_start"] and info["run_end"]:
        try:
            start = datetime.strptime(info["run_start"], "%Y-%m-%d %H:%M:%S")
            end = datetime.strptime(info["run_end"], "%Y-%m-%d %H:%M:%S")
            info["run_duration_min"] = round((end - start).total_seconds() / 60, 1)
        except ValueError:
            pass

    # If no explicit termination was found, infer from final state
    if info["termination_type"] is None:
        if info["final_context_pct"] and info["final_context_pct"] >= 90:
            info["termination_type"] = "context_window"
            info["termination_reason"] = "context window exhausted"
        elif info["max_failures"] >= 3:
            info["termination_type"] = "max_retries"
            info["termination_reason"] = "consecutive failures"
        elif info["max_no_progress"] >= 5:
            info["termination_type"] = "max_no_progress"
            info["termination_reason"] = "no progress limit"
        else:
            info["termination_type"] = "max_iterations"
            info["termination_reason"] = "iteration/API call limit"

    return info


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------

def collect_data():
    """Collect coverage data and run metadata from all runs."""
    data = {}
    for design_label, design_prefix in DESIGNS.items():
        data[design_label] = {}
        run_dirs = sorted(glob.glob(os.path.join(WORK_DIR, f"{design_prefix}_run_*")))
        run_nums = []
        for rd in run_dirs:
            m = re.search(r"_run_(\d+)$", rd)
            if m and os.path.isdir(rd):
                run_nums.append(int(m.group(1)))
        run_nums.sort()

        for run_num in run_nums:
            run_dir = os.path.join(WORK_DIR, f"{design_prefix}_run_{run_num}")
            cov_dir = os.path.join(run_dir, "coverage")

            # Coverage iteration reports
            iter_txts = sorted(glob.glob(os.path.join(cov_dir, "functional_coverage_iter_*.txt")))
            iter_nums = []
            for txt_path in iter_txts:
                m = re.search(r"functional_coverage_iter_(\d+)\.txt", txt_path)
                if m:
                    iter_nums.append(int(m.group(1)))
            iter_nums.sort()

            iter_xml_map = {}
            for xp in glob.glob(os.path.join(cov_dir, "iter_*_report.xml")):
                m = re.search(r"iter_(\d+)_report\.xml", xp)
                if m:
                    iter_xml_map[int(m.group(1))] = xp

            iterations = {}
            for iter_num in iter_nums:
                txt_path = os.path.join(cov_dir, f"functional_coverage_iter_{iter_num}.txt")
                funcov = parse_text_summary(txt_path)
                cg_pcts = parse_text_covergroup_types_multiline(txt_path)
                stmt = None
                if iter_num in iter_xml_map:
                    stmt = parse_xml_stmt_coverage(iter_xml_map[iter_num])
                iterations[iter_num] = {
                    "funcov_pct": funcov,
                    "stmt_pct": stmt,
                    "cg_pcts": cg_pcts,
                }

            cum_txt = os.path.join(cov_dir, "cumulative_functional_coverage.txt")
            cum_funcov = None
            cum_cg_pcts = {}
            if os.path.exists(cum_txt):
                cum_funcov = parse_text_summary(cum_txt)
                cum_cg_pcts = parse_text_covergroup_types_multiline(cum_txt)

            cum_xml = os.path.join(cov_dir, "cumulative_report.xml")
            cum_stmt = None
            if os.path.exists(cum_xml):
                cum_stmt = parse_xml_stmt_coverage(cum_xml)

            # Parse run log
            log_path = os.path.join(run_dir, "run.log")
            run_info = parse_run_log(log_path)

            data[design_label][run_num] = {
                "iterations": iterations,
                "iter_nums": iter_nums,
                "cumulative": {
                    "funcov_pct": cum_funcov,
                    "stmt_pct": cum_stmt,
                    "cg_pcts": cum_cg_pcts,
                },
                "run_info": run_info,
            }
    return data


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def create_excel(data, output_path):
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    cum_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    info_header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    pct_format = "0.00%"
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header(ws, row, c1, c2, fill=None):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = fill or header_fill
            cell.alignment = wrap_center
            cell.border = thin_border

    def style_subheader(ws, row, c1, c2):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center
            cell.border = thin_border

    def write_pct(ws, row, col, val, bold=False, fill=None):
        cell = ws.cell(row=row, column=col)
        if val is not None:
            cell.value = val / 100.0
            cell.number_format = pct_format
        else:
            cell.value = "N/A"
        cell.alignment = center
        cell.border = thin_border
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill

    def write_text(ws, row, col, val, bold=False, fill=None):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.alignment = center
        cell.border = thin_border
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill

    def write_num(ws, row, col, val, fmt=None, bold=False, fill=None):
        cell = ws.cell(row=row, column=col)
        cell.value = val
        cell.alignment = center
        cell.border = thin_border
        if fmt:
            cell.number_format = fmt
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill

    # Max iterations across all runs
    max_iters = 0
    for dl in data:
        for rn in data[dl]:
            n = len(data[dl][rn]["iter_nums"])
            max_iters = max(max_iters, n)

    # ===== Sheet 1: Functional Coverage (Covergroup %) =====
    ws = wb.active
    ws.title = "Functional Coverage"

    headers = ["Design", "Run"]
    for i in range(1, max_iters + 1):
        headers.append(f"Iter {i}")
    headers.append("Final\n(merged)")
    max_col = len(headers)

    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, 1, max_col)
    ws.row_dimensions[1].height = 30

    row = 2
    for design_label in DESIGNS:
        if design_label not in data:
            continue
        for run_num in sorted(data[design_label]):
            rd = data[design_label][run_num]
            iters = rd["iterations"]
            iter_nums = rd["iter_nums"]
            cum = rd["cumulative"]
            write_text(ws, row, 1, design_label)
            write_text(ws, row, 2, f"run_{run_num}")
            for idx, iter_num in enumerate(iter_nums):
                write_pct(ws, row, 3 + idx, iters[iter_num]["funcov_pct"])
            for i in range(len(iter_nums), max_iters):
                write_text(ws, row, 3 + i, "-")
            write_pct(ws, row, 3 + max_iters, cum["funcov_pct"], bold=True, fill=cum_fill)
            row += 1

    ws.column_dimensions[get_column_letter(1)].width = 20
    ws.column_dimensions[get_column_letter(2)].width = 10
    for c in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ===== Sheet 2: Code Coverage (Statement %) =====
    ws = wb.create_sheet(title="Code Coverage")
    headers2 = ["Design", "Run"]
    for i in range(1, max_iters + 1):
        headers2.append(f"Iter {i}")
    headers2.append("Final\n(merged)")
    max_col2 = len(headers2)
    for ci, h in enumerate(headers2, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, 1, max_col2)
    ws.row_dimensions[1].height = 30

    row = 2
    for design_label in DESIGNS:
        if design_label not in data:
            continue
        for run_num in sorted(data[design_label]):
            rd = data[design_label][run_num]
            iters = rd["iterations"]
            iter_nums = rd["iter_nums"]
            cum = rd["cumulative"]
            write_text(ws, row, 1, design_label)
            write_text(ws, row, 2, f"run_{run_num}")
            for idx, iter_num in enumerate(iter_nums):
                write_pct(ws, row, 3 + idx, iters[iter_num].get("stmt_pct"))
            for i in range(len(iter_nums), max_iters):
                write_text(ws, row, 3 + i, "-")
            write_pct(ws, row, 3 + max_iters, cum.get("stmt_pct"), bold=True, fill=cum_fill)
            row += 1

    ws.column_dimensions[get_column_letter(1)].width = 20
    ws.column_dimensions[get_column_letter(2)].width = 10
    for c in range(3, max_col2 + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # ===== Sheet 3: Run Info (config + runtime metadata) =====
    ws = wb.create_sheet(title="Run Info")

    # Column groups:
    #   Identity: Design, Run
    #   Coverage Result: Final Funcov %, Final Stmt %
    #   Runtime: API Calls, Iterations, Duration (min)
    #   Token Usage: Final Tokens, Context Used %, Context Window
    #   Termination: How Ended, Reason
    #   Config/Features: Driver Pipeline, Driver File,
    #                    Infra Mod Triggered, Infra Mod @ Iter, Infra Mod @ API Call
    #   Observed Limits: Max Failures, Max No-Progress
    info_headers = [
        # Identity
        "Design",
        "Run",
        # Coverage result
        "Final\nFuncov %",
        "Final\nStmt %",
        # Runtime
        "API Calls",
        "Iterations",
        "Duration\n(min)",
        # Token usage
        "Final\nTokens",
        "Context\nUsed %",
        "Context\nWindow",
        # Termination
        "Termination\nType",
        "Termination\nReason",
        # Features
        "Driver\nPipeline",
        "Driver\nFile",
        "Infra Mod\nTriggered",
        "Infra Mod\n@ Iter",
        "Infra Mod\n@ API Call",
        # Observed limits
        "Max\nFailures",
        "Max\nNo-Progress",
        # Timestamps
        "Run Start",
        "Run End",
    ]
    max_col3 = len(info_headers)

    for ci, h in enumerate(info_headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, 1, max_col3, fill=info_header_fill)
    ws.row_dimensions[1].height = 35

    row = 2
    for design_label in DESIGNS:
        if design_label not in data:
            continue
        for run_num in sorted(data[design_label]):
            rd = data[design_label][run_num]
            cum = rd["cumulative"]
            ri = rd["run_info"]
            c = 1

            # Identity
            write_text(ws, row, c, design_label); c += 1
            write_text(ws, row, c, f"run_{run_num}"); c += 1

            # Coverage result
            write_pct(ws, row, c, cum.get("funcov_pct")); c += 1
            write_pct(ws, row, c, cum.get("stmt_pct")); c += 1

            # Runtime
            write_num(ws, row, c, ri["total_api_calls"]); c += 1
            write_num(ws, row, c, ri["total_iterations"]); c += 1
            write_num(ws, row, c, ri["run_duration_min"], fmt="0.0"); c += 1

            # Token usage
            write_num(ws, row, c, ri["final_tokens"], fmt="#,##0"); c += 1
            if ri["final_context_pct"] is not None:
                write_pct(ws, row, c, ri["final_context_pct"])
            else:
                write_text(ws, row, c, "N/A")
            c += 1
            write_num(ws, row, c, ri["context_window"], fmt="#,##0"); c += 1

            # Termination
            write_text(ws, row, c, ri["termination_type"]); c += 1
            write_text(ws, row, c, ri["termination_reason"]); c += 1

            # Features
            write_text(ws, row, c, "Yes" if ri["driver_pipeline_enabled"] else "No"); c += 1
            write_text(ws, row, c, ri["driver_file"] or "-"); c += 1
            write_text(ws, row, c, "Yes" if ri["infra_modification_iter"] else "No"); c += 1
            write_text(ws, row, c, ri["infra_modification_iter"] or "-"); c += 1
            write_text(ws, row, c, ri["infra_modification_api_call"] or "-"); c += 1

            # Observed limits
            write_num(ws, row, c, ri["max_failures"]); c += 1
            write_num(ws, row, c, ri["max_no_progress"]); c += 1

            # Timestamps
            write_text(ws, row, c, ri["run_start"] or "-"); c += 1
            write_text(ws, row, c, ri["run_end"] or "-"); c += 1

            row += 1

    # Column widths for Run Info
    widths = {
        1: 18, 2: 10, 3: 12, 4: 12,
        5: 10, 6: 10, 7: 10,
        8: 12, 9: 11, 10: 12,
        11: 16, 12: 28,
        13: 10, 14: 22, 15: 10, 16: 10, 17: 12,
        18: 10, 19: 12,
        20: 20, 21: 20,
    }
    for col_num, w in widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = w

    # ===== Sheet 4+: Per-design detail with covergroup breakdown =====
    for design_label in DESIGNS:
        if design_label not in data:
            continue
        ws = wb.create_sheet(title=design_label[:31])

        row = 1
        for run_num in sorted(data[design_label]):
            rd = data[design_label][run_num]
            iters = rd["iterations"]
            iter_nums = rd["iter_nums"]
            cum = rd["cumulative"]
            n_iters = len(iter_nums)
            num_cols = 2 + n_iters

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            ws.cell(row=row, column=1, value=f"Run {run_num}")
            style_header(ws, row, 1, num_cols)
            row += 1

            write_text(ws, row, 1, "Metric")
            for idx in range(n_iters):
                write_text(ws, row, 2 + idx, f"Iter {idx + 1}")
            write_text(ws, row, 2 + n_iters, "Final (merged)")
            style_subheader(ws, row, 1, num_cols)
            row += 1

            write_text(ws, row, 1, "Covergroup %", bold=True)
            for idx, iter_num in enumerate(iter_nums):
                write_pct(ws, row, 2 + idx, iters[iter_num]["funcov_pct"])
            write_pct(ws, row, 2 + n_iters, cum["funcov_pct"], bold=True, fill=cum_fill)
            row += 1

            write_text(ws, row, 1, "Statement %", bold=True)
            for idx, iter_num in enumerate(iter_nums):
                write_pct(ws, row, 2 + idx, iters[iter_num].get("stmt_pct"))
            write_pct(ws, row, 2 + n_iters, cum.get("stmt_pct"), bold=True, fill=cum_fill)
            row += 1

            all_cg = sorted(cum.get("cg_pcts", {}).keys())
            for cg_name in all_cg:
                write_text(ws, row, 1, f"  {cg_name}")
                for idx, iter_num in enumerate(iter_nums):
                    pct = iters[iter_num].get("cg_pcts", {}).get(cg_name)
                    write_pct(ws, row, 2 + idx, pct)
                cum_pct = cum.get("cg_pcts", {}).get(cg_name)
                write_pct(ws, row, 2 + n_iters, cum_pct, bold=True, fill=cum_fill)
                row += 1

            row += 1

        ws.column_dimensions[get_column_letter(1)].width = 28
        for col in range(2, 15):
            ws.column_dimensions[get_column_letter(col)].width = 16

    output_path = os.path.abspath(output_path)
    wb.save(output_path)
    print(f"Excel saved to: {output_path}")


if __name__ == "__main__":
    data = collect_data()
    for design in data:
        for run in sorted(data[design]):
            rd = data[design][run]
            iters = rd["iterations"]
            iter_nums = rd["iter_nums"]
            cum = rd["cumulative"]
            ri = rd["run_info"]
            parts = []
            for n in iter_nums:
                v = iters[n]["funcov_pct"]
                parts.append(f"{v:.1f}%" if v is not None else "N/A")
            progression = " → ".join(parts)
            final = cum["funcov_pct"]
            print(
                f"{design} run_{run}: "
                f"[{progression}] → Final {final}% | "
                f"API={ri['total_api_calls']} Iters={ri['total_iterations']} "
                f"Tokens={ri['final_tokens']} ({ri['final_context_pct']}%) "
                f"End={ri['termination_type']} "
                f"Driver={'Y' if ri['driver_pipeline_enabled'] else 'N'}"
            )
    create_excel(data, "coverage_report.xlsx")
